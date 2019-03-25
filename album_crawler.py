# coding: utf-8

import os
import json
import time
import crawler as craw

from progressbar import *
from logger import Logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

''' 专辑 '''
album_id = input("请输入专辑的id: ")

def my_exit():
    temp = input("按Enter键退出......")
    exit()

'''尝试获取专辑的信息'''
try:
    album_info_dic = craw.GetAlbumInfo(album_id)
except Exception as identifier:
    print('专辑信息获取失败，错误信息如下：')
    print(identifier)
    my_exit()

album_name = album_info_dic['name']
album_name_bk = album_info_dic['name']
album_song = album_info_dic['songs']
album_cnt_comment = album_info_dic['cnt_comment']

__folder_path = os.getcwd() + '\\Albums'


''' 专辑配置文件'''
has_configuration = False        # 标志是否有配置文件 默认为 False
ini_info = {}                    # 保存配置信息的字典
ini_file_path = __folder_path + '\\' + album_id + '.json'  # 配置文件路径
log_file_path = __folder_path + '\\' + album_id + '.log'    # 日志文件路径


''' 更新专辑配置文件'''
def UpdateIniFile():
    json_str = json.dumps(ini_info, indent=4, ensure_ascii=False)
    with open(ini_file_path, 'w', encoding='utf-8') as json_file:
        json_file.write(json_str)


''' 尝试在当前路径创建 Albums 文件夹 '''
try:
    os.mkdir(__folder_path)
except FileExistsError: # 如果Albums文件夹已存在，则尝试加载配置文件
    logger = Logger(level_s=Logger.INFO, level_f=Logger.INFO, filename=log_file_path).getlogger()
    logger.info('文件夹 "%s" 已存在.' % (__folder_path))

    try:  # 尝试从配置文件读取配置信息
        with open(ini_file_path, 'r', encoding='utf-8') as json_file:
            ini_info = json.load(json_file)
    except Exception as identifier:  # 配置文件不存在或者配置信息加载失败
        logger.info('未检测到配置文件：%s' % (identifier))
    else:
        has_configuration = True
except Exception as identifier: # 如果无法创建Albums文件夹，则退出
    print('[ ERROR ]: 无法在当前路径创建Albums文件夹')
    my_exit()
else:  # 如果 Albums文件夹创建成功
    logger = Logger(level_s=Logger.INFO, level_f=Logger.INFO, filename=log_file_path).getlogger()
    logger.info('成功创建文件夹 "%s"' % (__folder_path))


count = 1
page = 1
cnt_comment = album_cnt_comment

album_no_record = True
crawer_not_complete = True
folder_not_exist = False
logger.name = '《' + album_name_bk + '》'


if has_configuration == False: # 如果没有发现配置文件
    folder_path = __folder_path + '\\' + album_name # 尝试使用爬取到的专辑名作为文件夹名
    try: # 检测该文件名是否符合规范
        os.mkdir(folder_path)
    except FileExistsError:
        pass
    except Exception as identifier:
        folder_not_exist = True
        logger.error(" '%s' 不是有效的文件夹名，使用其id:'%s'作为文件夹名： " %(album_name, album_id))
        album_name = album_id
        folder_path = __folder_path + '\\' + album_name
    else:
        logger.info('成功创建文件夹 "%s"' % (folder_path))
elif album_id in ini_info: # 如果存在配置文件 且 文件内容符合规范
    item = ini_info[album_id]

    if item['State']:  # 如果专辑状态为 True，则表示已爬取或忽略
        crawer_not_complete = False          # 专辑评论已爬取完毕，记为 False
        logger.info('历史信息记录该专辑的评论已爬取完毕，跳过...')
    elif item['TotalComment'] > 0:  # 专辑状态为 False 且其他信息不是初始状态
        page = item['NextPage']              # 更新下一个要爬取的评论所在页码
        count = 20*(page-1)+1                # 更新已爬取评论数目
        cnt_comment = item['TotalComment']   # 更新总的评论数目
        album_no_record = False              # 已经获取了专辑信息 所以置为 False

    album_name = item['FileName']        # 更新 Excel 文件名
    folder_path = __folder_path + '\\' + album_name


''' Excel文件格式相关'''
excel_header = ['null', 'commentId', 'content',
                'likedCount', 'parentCommentId', 'time', 'userId']
column_width = {'A': 15, 'B': 100, 'C': 12, 'D': 15, 'E': 20, 'F': 15}
alignment = Alignment(horizontal='center', vertical='center')
time_format = "%Y-%m-%d %H:%M:%S"

if crawer_not_complete:

    if folder_not_exist:
        try:
            os.mkdir(folder_path)  # folder_path = os.getcwd() + '\\Albums' + album_id
        except Exception as identifier:  # 如果 Albums文件夹 创建失败，则退出程序
            logger.error('无法创建文件夹：%s' % (identifier))
            my_exit()
        else:  # 如果 Albums文件夹创建成功
            logger.info('成功创建文件夹 "%s"' % (folder_path))

    excel_file_path = folder_path + '\\《' + album_name + '》.xlsx'

    if album_no_record:
        ini_info[album_id] = {'AlbumName': album_name_bk, 'State': False, 'NextPage': 1, 'GetedComment': 0, 'TotalComment': cnt_comment, 'FileName': album_name}
        UpdateIniFile()

        ini_info['songs'] = {}
        for song in album_song:
            song_id = album_song[song]
            ini_info['songs'][song_id] ={'SongName': song, 'State': False, 'NextPage': 1, 'GetedComment': 0, 'TotalComment': -1, 'FileName':''}
        UpdateIniFile()

        logger.info('查询到 %d 条评论...' % (cnt_comment))

        wb = Workbook()
        ws = wb.active

        for i in range(1, 7):  # 设置表头
            temp = ws.cell(1, i, excel_header[i])
            temp.alignment = alignment
        for column in column_width:  # 设置列宽
            ws.column_dimensions[column].width = column_width[column]

        wb.save(excel_file_path)
    else:
        wb = load_workbook(excel_file_path)                  # 加载 Excel文件
        ws = wb.active                                       # 获取工作表
        logger.info('共 %d 条评论，已爬取 %d 条，Excel文件名为：%s' %(cnt_comment, count-1, item['FileName']))

    # 创建进度条
    processbar_title = '正在爬取《'+album_name+'》的评论: '
    widgets = [processbar_title, Percentage(), '', Bar('>'), '', ETA()]
    pbar = ProgressBar(cnt_comment, widgets).start()

    # 获取专辑的所有评论，若爬取过程中出错，则保存已爬取的信息，然后跳过该专辑
    all_success = True
    max_page = cnt_comment//20 + 1
    while count <= cnt_comment and page <= max_page:
        # 爬取第page页的评论
        try:
            json_comment = craw.GetAlbumComments(album_id, page)
        except Exception as identifier:
            all_success = False
            logger.error('第%d页的评论爬取失败...' % (page))
            break

        if json_comment['code'] != 200:
            all_success = False
            logger.error('获取评论失败，错误代码：%d, 错误信息：%s'%(json_comment['code'], json_comment['msg']))
            break
        # 将该页的评论写入excel
        comments = json_comment['comments']
        for comment in comments:
            count = count + 1
            ws.cell(count, 1, comment['commentId'])
            content = comment['content'].split()
            content = " ".join(content)
            try:
                ws.cell(count, 2, content)
            except Exception as identifier:
                logger.warn('无法将第 %d 条评论 “%s” 写入到Excel' % (count-1, content))

            ws.cell(count, 3, comment['likedCount'])
            ws.cell(count, 4, comment['parentCommentId'])

            timeArray = time.localtime(comment['time']/1000)
            time_str = time.strftime(time_format, timeArray)
            ws.cell(count, 5, time_str)

            ws.cell(count, 6, comment['user']['userId'])

        pbar.update(count)  # 更新进度条

        # 每隔5页保存一次Excel，同时更新配置文件
        if page % 10 == 0:
            wb.save(excel_file_path)
            ini_info[album_id]['NextPage'] = page
            ini_info[album_id]['GetedComment'] = count-1
            UpdateIniFile()

        page = page + 1

    wb.save(excel_file_path)
    ini_info[album_id]['NextPage'] = page
    ini_info[album_id]['GetedComment'] = count-1

    if all_success:
        ini_info[album_id]['State'] = True
        logger.info('爬取完毕！实际共爬取了 %d 条评论...' % (count-1))

    UpdateIniFile()




''' 爬取该专辑内所有歌曲的评论 '''
logger.name = 'root'
logger.info('开始爬取专辑内歌曲的评论，共查询到 %d 首歌曲' % len(album_song))

song_record_dic = ini_info['songs']

for song_name in album_song:

    count = 1
    page = 1
    song_id = album_song[song_name]
    song_not_record = True
    logger.name = '《'+song_name+'》'
    excel_file_path = folder_path + '\\' + song_name + '.xlsx' # 首先将歌曲名作为 Excel文件名

    if song_id in song_record_dic:
        item = song_record_dic[song_id]
        if item['State']:  # 如果状态为 True，则表示已爬取或忽略
            logger.info('历史信息记录该歌曲的评论已爬取完毕，跳过...')
            continue
        elif item['TotalComment'] > 0:  # 专辑状态为 False 且其他信息不是初始状态
            page = item['NextPage']              # 更新下一个要爬取的评论所在页码
            count = 20*(page-1)+1                # 更新已爬取评论数目
            cnt_comment = item['TotalComment']   # 更新总的评论数目
            song_not_record = False              # 已经获取了专辑信息 所以置为 False
            excel_file_path = folder_path + '\\' + item['FileName']        # 更新 Excel 文件名
            wb = load_workbook(excel_file_path)                  # 加载 Excel文件
            ws = wb.active                                       # 获取工作表
            logger.info('共 %d 条评论，已爬取 %d 条，Excel文件名为：%s' %
                        (cnt_comment, count-1, item['FileName']))

    if song_not_record:
        try:
            page1_comment = craw.GetSongComments(song_id, 1)
            cnt_comment = page1_comment['total']
        except Exception as identifier:
            logger.error('信息获取失败：%s' % (identifier))
            continue

        ini_info['songs'][song_id] ={'SongName': song_name, 'State': False, 'NextPage': 1, 'GetedComment': 0, 'TotalComment': cnt_comment, 'FileName':''}

        logger.info('信息获取完毕，共查询到 %d 条评论...' % (cnt_comment))

        # 创建 excel
        wb = Workbook()
        ws = wb.active

        for i in range(1, 7):  # 设置表头
            temp = ws.cell(1, i, excel_header[i])
            temp.alignment = alignment
        for column in column_width:  # 设置列宽
            ws.column_dimensions[column].width = column_width[column]

        try:  # 尝试保存表格，检查文件名是否符合标准
            wb.save(excel_file_path)
        except Exception as identifier:
            excel_file_path = folder_path + '\\' + song_id + '.xlsx'
            wb.save(excel_file_path)
            logger.error('"%s"不是合法的文件名，使用其id:"%s"作为新文件名.' %
                         (song_name, song_id))
            ini_info['songs'][song_id]['FileName'] = song_id + '.xlsx'
        else:
            ini_info['songs'][song_id]['FileName'] = song_name+'.xlsx'

        UpdateIniFile()

    ''' 创建进度条'''
    processbar_title = '正在爬取《'+song_name+'》的评论: '
    widgets = [processbar_title, Percentage(), '', Bar('>'), '', ETA()]
    pbar = ProgressBar(cnt_comment, widgets).start()

    ''' 获取专辑的所有评论，若爬取过程中出错，则保存已爬取的信息，然后跳过该专辑 '''
    all_success = True
    max_page = cnt_comment//20 + 1
    while count <= cnt_comment and page <= max_page:

        ''' 爬取第page页的评论'''
        try:
            json_comment = craw.GetSongComments(song_id, page)
        except Exception as identifier:
            all_success = False
            logger.error('第%d页的评论爬取失败...' % (page))
            break
        if json_comment['code'] != 200:
            all_success = False
            logger.error('获取评论失败，错误代码：%d, 错误信息：%s'%(json_comment['code'], json_comment['msg']))
            break

        ''' 将该页的评论写入excel '''
        comments = json_comment['comments']
        for comment in comments:
            count = count + 1

            ws.cell(count, 1, comment['commentId'])

            content = comment['content'].split()
            content = " ".join(content)
            try:
                ws.cell(count, 2, content)
            except Exception as identifier:
                logger.warn('无法将第 %d 条评论 “%s” 写入到Excel' % (count-1, content))

            ws.cell(count, 3, comment['likedCount'])
            ws.cell(count, 4, comment['parentCommentId'])

            timeArray = time.localtime(comment['time']/1000)
            time_str = time.strftime(time_format, timeArray)
            ws.cell(count, 5, time_str)

            ws.cell(count, 6, comment['user']['userId'])

        pbar.update(count)  # 更新进度条

        # 每隔5页保存一次Excel，同时更新配置文件
        if page % 10 == 0:
            wb.save(excel_file_path)
            ini_info['songs'][song_id]['NextPage'] = page
            ini_info['songs'][song_id]['GetedComment'] = count-1
            UpdateIniFile()

        page = page + 1

    wb.save(excel_file_path)
    ini_info['songs'][song_id]['NextPage'] = page
    ini_info['songs'][song_id]['GetedComment'] = count-1
    if all_success:
        ini_info['songs'][song_id]['State'] = True
        logger.info('爬取完毕！实际共爬取了 %d 条评论...' % (count-1))
    UpdateIniFile()

logger.name = 'root'
logger.info('>>>>>>>> End！ <<<<<<<<')
my_exit()
