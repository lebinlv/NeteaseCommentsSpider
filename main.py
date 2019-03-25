# coding: utf-8

import os
import json
import time
import crawler as craw

from progressbar import *
from logger import Logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

''' 歌手相关 '''
singer = input("请输入歌手姓名：")
singer_id = input("请输入该歌手的id：")


''' Excel 格式相关 '''
excel_header = ['null', 'commentId', 'content',
                'likedCount', 'parentCommentId', 'time', 'userId']
column_width = {'A': 15, 'B': 100, 'C': 12, 'D': 15, 'E': 20, 'F': 15}
alignment = Alignment(horizontal='center', vertical='center')
time_format = "%Y-%m-%d %H:%M:%S"


folder_path = os.getcwd() + '\\' + singer


''' 配置文件相关 '''
has_configuration = False        # 标志是否有配置文件 默认为 False
ini_info = {}                    # 保存配置信息的字典
ini_file_path = folder_path + '\\'+singer+'.ini'  # 配置文件路径


def UpdateConfigurationFile():
    ''' 更新配置文件 '''
    json_str = json.dumps(ini_info, indent=4, ensure_ascii=False)
    with open(ini_file_path, 'w', encoding='utf-8') as json_file:
        json_file.write(json_str)


''' 某些专辑名无法作为文件名 '''
excel_name_error = 0


''' 尝试创建保存excel评论文件的文件夹 '''
try:
    os.mkdir(folder_path)
except FileExistsError:  # 如果文件夹已存在，则尝试加载配置文件
    logger = Logger(level_s=Logger.INFO, level_f=Logger.INFO,
                    filename=folder_path+'\\'+singer+'.log').getlogger()
    logger.info('文件夹 "%s" 已存在...' % (folder_path))
    ''' 尝试从配置文件读取配置信息 '''
    try:
        with open(ini_file_path, 'r', encoding='utf-8') as json_file:
            ini_info = json.load(json_file)
    except Exception as identifier:  # 配置文件不存在或者配置信息加载失败
        has_configuration = False
        logger.info('加载配置文件失败：%s' % (identifier))
    else:
        has_configuration = True
except Exception as identifier:
    print('文件夹创建失败：%s' % (identifier))
    exit()
else:
    logger = Logger(level_s=Logger.INFO, level_f=Logger.INFO,
                    filename=folder_path+'\\'+singer+'.log').getlogger()
    logger.info('创建文件夹：%s' % (folder_path))

folder_path = folder_path + '\\'


''' 获取歌手的专辑列表 '''
try:
    album_dic = craw.GetAlbumId(singer_id)
except Exception as identifier:
    logger.error('无法获取 %s 的专辑列表：%s' % (singer, identifier))
    exit()
else:
    logger.info('专辑列表获取完毕，共查询到 %d 张专辑...' % (len(album_dic)))


''' 更新配置信息，并将其写入文件 '''
if not has_configuration:
    for album_name in album_dic:
        ini_info[album_name] = {'state': False, 'pPage': 1, 'getedComment': 0,
                                'totalComment': -1, 'albumId': '', 'fileName': ''}
    UpdateConfigurationFile()

def GetAlbumComments(album_dic, logger)
    ''' 依次爬取各专辑的评论信息 '''
    for album_name in album_dic:

        count = 1
        page = 1
        album_id = album_dic[album_name]
        album_info_empty = True
        logger.name = '《' + album_name + '》'

        excel_file_path = folder_path+album_name+'.xlsx'  # 首先将专辑名作为 Excel文件名

        if has_configuration and album_name in ini_info:
            item = ini_info[album_name]
            if item['state']:                        # 如果专辑状态为 True，则表示已爬取或忽略，
                logger.info('配置文件记录该专辑已爬取完毕，跳过...')
                continue
            elif item['totalComment'] > 0:             # 专辑状态为 False 且其他信息不是初始状态
                page = item['pPage']                 # 更新下一个要爬取的评论所在页码
                count = 20*(page-1)+1                # 更新已爬取评论数目
                cnt_comment = item['totalComment']   # 更新总的评论数目
                album_info_empty = False                             # 已经获取了专辑信息 所以置为 False
                excel_file_path = folder_path+item['fileName']       # 更新 Excel 文件名
                wb = load_workbook(excel_file_path)                  # 加载 Excel文件
                ws = wb.active                                       # 获取工作表
                logger.info('共 %d 条评论，已爬取 %d 条，Excel文件名为：%s' %
                            (cnt_comment, count-1, item['fileName']))

        # 如果没有找到专辑信息，则需要重新从网上爬取
        if album_info_empty:
            ''' 爬取专辑信息 '''
            try:
                song_list, cnt_comment = craw.GetAlbumInfo(album_id)
            except Exception as identifier:
                logger.error('信息获取失败：%s' % (identifier))
                continue

            if album_name in ini_info:
                ini_info[album_name]['albumId'] = album_id
                ini_info[album_name]['totalComment'] = cnt_comment
            else:
                ini_info[album_name] = {'state': False, 'pPage': 1, 'getedComment': 0,
                                        'totalComment': cnt_comment, 'albumId': album_id, 'fileName': ''}

            logger.info('信息获取完毕，共查询到 %d 条评论...' % (cnt_comment))

            # 创建 excel
            wb = Workbook()
            ws = wb.active
            # 设置表头
            for i in range(1, 7):
                temp = ws.cell(1, i, excel_header[i])
                temp.alignment = alignment
            # 设置列宽
            for column in column_width:
                ws.column_dimensions[column].width = column_width[column]
            # 尝试保存表格，检查文件名是否符合标准
            try:
                wb.save(excel_file_path)
            except Exception as identifier:
                excel_name_error = excel_name_error + 1
                new_name = 'ErrorName'+str(excel_name_error)+'.xlsx'
                excel_file_path = folder_path+new_name
                wb.save(excel_file_path)
                logger.error('"%s"保存失败，已另存为 "%s". %s' %
                            (album_name+'.xlsx', new_name, identifier))
                ini_info[album_name]['fileName'] = new_name
            else:
                ini_info[album_name]['fileName'] = album_name+'.xlsx'

            UpdateConfigurationFile()

        ''' 创建进度条'''
        processbar_title = '正在爬取《'+album_name+'》的评论: '
        widgets = [processbar_title, Percentage(), '', Bar('>'), '', ETA()]
        pbar = ProgressBar(cnt_comment, widgets).start()

        ''' 获取专辑的所有评论，若爬取过程中出错，则保存已爬取的信息，然后跳过该专辑 '''
        all_success = True
        max_page = cnt_comment//20 + 1
        while count <= cnt_comment and page <= max_page:

            ''' 爬取第page页的评论'''
            try:
                json_comment = craw.GetAlbumComments(album_id, page)
            except Exception as identifier:
                all_success = False
                logger.error('第%d页的评论爬取失败...' % (page))
                break

            ''' 将该页的评论写入excel '''
            comments = json_comment['comments']
            for comment in comments:
                count = count + 1

                ws.cell(count, 1, comment['commentId'])

                content = comment['content'].split()
                content = " ".join(content)
                try:
                    ws.cell(count, 2, content)
                except Exception as identifier:
                    logger.warn('无法将第 %d 条评论 “%s” 写入到Excel' % (count-1, content))

                ws.cell(count, 3, comment['likedCount'])
                ws.cell(count, 4, comment['parentCommentId'])

                timeArray = time.localtime(comment['time']/1000)
                time_str = time.strftime(time_format, timeArray)
                ws.cell(count, 5, time_str)

                ws.cell(count, 6, comment['user']['userId'])

            pbar.update(count)  # 更新进度条

            # 每隔5页保存一次Excel，同时更新配置文件
            if page % 10 == 0:
                wb.save(excel_file_path)
                ini_info[album_name]['pPage'] = page
                ini_info[album_name]['getedComment'] = count-1
                UpdateConfigurationFile()

            page = page + 1

        wb.save(excel_file_path)
        ini_info[album_name]['pPage'] = page
        ini_info[album_name]['getedComment'] = count-1
        if all_success:
            ini_info[album_name]['state'] = True
            logger.info('爬取完毕！共爬取了 %d 条评论...' % (count-1))
        UpdateConfigurationFile()

logger.name = 'root'
logger.info('该歌手的所有专辑已爬取完毕！')
